import os
import active
import datetime
from os import listdir
from openpyxl import Workbook

def ensembleToSpreadsheet(filename, col, ws):
    arr = []
    for line in open(filename, 'r'):
        arr.append(float(line))
    for x in range(0, len(arr)):
        ws[col + str(x+2)] = arr[x]

# Create spreadsheets with data filled in so I only have to select data and graph it.
wb = Workbook()
ws = wb.active

ws.title = "Data"

# Create labels for each column
ws['A1'] = "DJIA"
ws['B1'] = "ENS_AVG"
ws['C1'] = "ENS_MED"
ws['D1'] = "ENS_MIN"
ws['E1'] = "ENS_MAX"
ws['F1'] = "AVG-DJIA"
ws['G1'] = "MED-DJIA"
ws['H1'] = "AVG-STDFIT"
ws['I1'] = "MED-STDFIT"

# Get the abs difference between expected and actual value
for x in range(2, 80):
    ws['F' + str(x)] = "=ABS(B" + str(x) + "-A" + str(x) + ")^2"
    ws['G' + str(x)] = "=ABS(C" + str(x) + "-A" + str(x) + ")^2"  

# Sum of Squared Residuals for Ensembles
ws['H2'] = "=SUM(F2:F79)" # Avg
ws['I2'] = "=SUM(G2:G79)" # Median

activeDirectory = active.listActiveDirectory()
files = listdir(activeDirectory)

input = "D:/Development/GitHub/GPP/GeneticProgrammingPortfolio/input/DJIA/djia_close_12-16_testing_90.txt"
input_arr = []

# Read and write Dow Jones Index
for line in open(input, 'r'):
    input_arr.append(float(line))
for x in range(0, len(input_arr)):
    ws['A' + str(x+2)] = input_arr[x]

# All ensemble statistics
ensembleToSpreadsheet(activeDirectory + "/ensemble-avg.txt", 'B', ws)
ensembleToSpreadsheet(activeDirectory + "/ensemble-med.txt", 'C', ws)
ensembleToSpreadsheet(activeDirectory + "/ensemble-min.txt", 'D', ws)
ensembleToSpreadsheet(activeDirectory + "/ensemble-max.txt", 'E', ws)

# Print a confirmation for GPP output
print("'data.xlsx' successfully created!")

# Save the file
wb.save(activeDirectory + "/data.xlsx")

